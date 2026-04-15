from __future__ import annotations

import csv
import io
import logging
from collections import Counter
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import HTTPException, status
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from . import repository as repo
from . import schemas

logger = logging.getLogger(__name__)


# =============================================================================
# EXISTING — summary + admin
# =============================================================================

async def fetch_summary(
    db: AsyncSession,
    *,
    tenant_id: str,
) -> schemas.DashboardSummaryResponse:
    logger.debug("[dashboard] fetch_summary tenant=%s", tenant_id)
    summary = await repo.get_dashboard_summary(db, tenant_id=tenant_id)
    logger.debug("[dashboard] fetch_summary complete tenant=%s", tenant_id)
    return schemas.DashboardSummaryResponse(summary=summary)


async def fetch_admin_overview(db: AsyncSession) -> schemas.AdminOverviewResponse:
    logger.debug("[dashboard] fetch_admin_overview")
    row = await repo.get_admin_overview(db)

    plan_dist = [
        schemas.PlanDistribution(**p)
        for p in (row.get("plan_distribution") or [])
    ]

    daily = [
        schemas.DailyActivity(**d)
        for d in (row.get("daily_activity") or [])
    ]

    return schemas.AdminOverviewResponse(
        tenants=schemas.TenantStatusBreakdown(
            total=row.get("tenant_total", 0),
            active=row.get("tenant_active", 0),
            trial=row.get("tenant_trial", 0),
            pending_plan=row.get("tenant_pending_plan", 0),
            suspended=row.get("tenant_suspended", 0),
            cancelled=row.get("tenant_cancelled", 0),
        ),
        total_users=row.get("total_users", 0),
        total_conversations=row.get("total_conversations", 0),
        conversations_this_month=row.get("conversations_this_month", 0),
        active_conversations=row.get("active_conversations", 0),
        leads=schemas.LeadBreakdown(
            hot=row.get("lead_hot", 0),
            nurture=row.get("lead_nurture", 0),
            cold=row.get("lead_cold", 0),
            total=row.get("lead_total", 0),
        ),
        revenue=schemas.RevenueMetrics(
            mrr_aud=float(row.get("mrr_aud", 0)),
            arr_aud=float(row.get("arr_aud", 0)),
            active_subscriptions=row.get("active_subscriptions", 0),
            trialing_subscriptions=row.get("trialing_subscriptions", 0),
            past_due_subscriptions=row.get("past_due_subscriptions", 0),
            plan_distribution=plan_dist,
        ),
        chatbots=schemas.ChatbotStats(
            active=row.get("chatbot_active", 0),
            draft=row.get("chatbot_draft", 0),
            inactive=row.get("chatbot_inactive", 0),
        ),
        daily_activity=daily,
    )


async def fetch_admin_tenants(
    db: AsyncSession,
    *,
    limit: int,
    offset: int,
) -> schemas.AdminTenantsResponse:
    logger.debug("[dashboard] fetch_admin_tenants limit=%s offset=%s", limit, offset)
    rows, total = await repo.get_admin_tenants(db, limit=limit, offset=offset)
    return schemas.AdminTenantsResponse(
        tenants=[schemas.TenantRow(**r) for r in rows],
        total=total,
        limit=limit,
        offset=offset,
    )


# =============================================================================
# LEADS — list, detail, export
# =============================================================================

def _build_lead_detail_response(raw: dict) -> schemas.LeadDetailResponse:
    stats = schemas.ConversationStats(
        user_messages=raw.get("user_messages") or 0,
        assistant_messages=raw.get("assistant_messages") or 0,
        avg_response_ms=raw.get("avg_response_ms"),
        min_response_ms=raw.get("min_response_ms"),
        max_response_ms=raw.get("max_response_ms"),
        total_tokens_messages=raw.get("total_tokens_messages"),
        first_message_at=raw.get("first_message_at"),
        last_message_at=raw.get("last_message_at"),
        conversation_duration_sec=raw.get("conversation_duration_sec"),
    )

    messages = [
        schemas.MessageItem(
            public_id=m["public_id"],
            role=m["role"],
            content=m["content"],
            tokens_used=m.get("tokens_used"),
            model_used=m.get("model_used"),
            response_ms=m.get("response_ms"),
            created_at=m["created_at"],
        )
        for m in raw.get("messages", [])
    ]

    stat_keys = {
        "user_messages", "assistant_messages", "avg_response_ms",
        "min_response_ms", "max_response_ms", "total_tokens_messages",
        "first_message_at", "last_message_at", "conversation_duration_sec",
        "messages",
    }
    conv_data = {k: v for k, v in raw.items() if k not in stat_keys}

    return schemas.LeadDetailResponse(**conv_data, stats=stats, messages=messages)


def _flatten_row(row: dict) -> dict:
    """Flatten array fields to comma-separated strings for CSV/Excel export."""
    flat = {}
    for k, v in row.items():
        if isinstance(v, list):
            flat[k] = ", ".join(str(x) for x in v) if v else ""
        elif isinstance(v, datetime):
            flat[k] = v.isoformat()
        else:
            flat[k] = v
    return flat


async def fetch_leads(
    db: AsyncSession,
    *,
    tenant_id: str,
    page: int,
    page_size: int,
) -> schemas.LeadsListResponse:
    data = await repo.get_leads_paginated(
        db, tenant_id=tenant_id, page=page, page_size=page_size
    )
    leads = [schemas.LeadListItem(**row) for row in data["leads"]]
    pagination = schemas.PaginationMeta(**data["pagination"])
    return schemas.LeadsListResponse(leads=leads, pagination=pagination)


async def fetch_lead_detail(
    db: AsyncSession,
    *,
    tenant_id: str,
    public_id: str,
) -> schemas.LeadDetailResponse:
    raw = await repo.get_lead_detail(db, tenant_id=tenant_id, public_id=public_id)
    if not raw:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Conversation '{public_id}' not found.",
        )
    return _build_lead_detail_response(raw)


# =============================================================================
# EXPORT — shared style constants
# =============================================================================

C = {
    "brown":       "A47764",
    "brownShade":  "D9B8A8",
    "burntOrange": "CC5500",
    "teal":        "069494",
    "taupe":       "54463A",
    "light":       "F7F3F1",
    "white":       "FFFFFF",
    "dark":        "1E1E1E",
    "hot":         "D93025",
    "warm":        "0F7B6C",
    "nurture":     "E8780C",
    "cold":        "3A5F8A",
}

TIER_ACCENT = {"hot": C["hot"], "warm": C["warm"], "nurture": C["nurture"], "cold": C["cold"]}
SENT_FILL   = {"positive": C["teal"], "neutral": C["brownShade"], "negative": C["burntOrange"]}


def _color_pie_slices(chart, colors: list) -> None:
    from openpyxl.chart.series import DataPoint
    ser = chart.series[0]
    for idx, hex_color in enumerate(colors):
        dp = DataPoint(idx=idx)
        dp.spPr.solidFill = hex_color[-6:]
        ser.dPt.append(dp)


def _color_bar_series(chart, colors: list) -> None:
    from openpyxl.chart.series import DataPoint
    ser = chart.series[0]
    for idx, hex_color in enumerate(colors):
        dp = DataPoint(idx=idx)
        dp.spPr.solidFill = hex_color[-6:]
        ser.dPt.append(dp)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color)


def _font(color=C["dark"], bold=False, size=10, name="Arial") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def _border(color=C["brownShade"]) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP        = Alignment(vertical="center", wrap_text=True)
CELL_BORDER = _border(C["brownShade"])


def _fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d %b %Y %H:%M")
    s = str(val).strip()
    if not s or s in ("-", "--", "n/a", "none", "null"):
        return ""
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%d %b %Y %H:%M")
        except ValueError:
            continue
    return ""


# =============================================================================
# CSV EXPORT
# =============================================================================

async def export_leads_csv(
    db,
    *,
    tenant_id: str,
    date_from: Optional[str],
    date_to: Optional[str],
    last_days: Optional[int],
) -> bytes:
    rows = await repo.get_leads_for_export(
        db, tenant_id=tenant_id,
        date_from=date_from, date_to=date_to, last_days=last_days,
    )
    if not rows:
        rows = []

    flat_rows = [_flatten_row(r) for r in rows]

    raw_fields = [
        "conversation_id", "lead_name", "lead_email", "lead_phone",
        "total_messages", "status", "ai_summary", "last_activity_at", "created_at",
        "lead_tier", "lead_score", "intent", "budget_min", "budget_max",
        "budget_currency", "suburbs_mentioned", "cities_mentioned",
        "property_types", "bedrooms_wanted", "timeline", "sentiment",
        "engagement_score", "topics_mentioned",
    ]

    header_map = {
        "conversation_id":   "Conversation ID",
        "lead_name":         "Name",
        "lead_email":        "Email",
        "lead_phone":        "Phone Number",
        "total_messages":    "Messages",
        "status":            "Status",
        "ai_summary":        "Summary",
        "last_activity_at":  "Last Activity",
        "created_at":        "Created At",
        "lead_tier":         "Tier",
        "lead_score":        "Score",
        "intent":            "Intent",
        "budget_min":        "Budget Min",
        "budget_max":        "Budget Max",
        "budget_currency":   "Currency",
        "suburbs_mentioned": "Suburbs",
        "cities_mentioned":  "Cities",
        "property_types":    "Property Types",
        "bedrooms_wanted":   "Bedrooms Wanted",
        "timeline":          "Timeline",
        "sentiment":         "Sentiment",
        "engagement_score":  "Engagement",
        "topics_mentioned":  "Topics",
    }

    DATE_FIELDS = {"last_activity_at", "created_at"}
    friendly_headers = [header_map.get(f, f) for f in raw_fields]

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["STATUS GUIDE"])
    writer.writerow(["Status", "What it means"])
    writer.writerow(["in_progress", "Chat is active — visitor is currently messaging."])
    writer.writerow(["closed",      "Auto-closed after 15 min of inactivity by the hourly job."])
    writer.writerow(["summarized",  "AI has extracted lead insights from the closed conversation."])
    writer.writerow(["emailed",     "Summary sent to the visitor (if email provided) and internal team."])
    writer.writerow(["archived",    "Conversation is old and has been moved to long-term storage."])
    writer.writerow([])

    writer.writerow(friendly_headers)
    for row in flat_rows:
        writer.writerow([
            _fmt_date(row.get(f, "")) if f in DATE_FIELDS else row.get(f, "")
            for f in raw_fields
        ])

    return output.getvalue().encode("utf-8")


# =============================================================================
# XLSX EXPORT
# =============================================================================

async def export_leads_xlsx(
    db,
    *,
    tenant_id: str,
    date_from: Optional[str],
    date_to: Optional[str],
    last_days: Optional[int],
) -> bytes:
    rows = await repo.get_leads_for_export(
        db, tenant_id=tenant_id,
        date_from=date_from, date_to=date_to, last_days=last_days,
    )
    flat_rows = [_flatten_row(r) for r in rows]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Leads Data ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leads"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    fieldnames = [
        "conversation_id", "lead_name", "lead_email", "lead_phone",
        "total_messages", "status", "lead_tier", "lead_score", "intent",
        "budget_min", "budget_max", "budget_currency", "suburbs_mentioned",
        "property_types", "timeline", "sentiment", "engagement_score",
        "ai_summary", "last_activity_at", "created_at",
    ]

    pretty_headers = {
        "conversation_id":   "Conversation ID",
        "lead_name":         "Name",
        "lead_email":        "Email",
        "lead_phone":        "Phone Number",
        "total_messages":    "Messages",
        "status":            "Status",
        "lead_tier":         "Tier",
        "lead_score":        "Score",
        "intent":            "AI Intent",
        "budget_min":        "Budget Min",
        "budget_max":        "Budget Max",
        "budget_currency":   "Currency",
        "suburbs_mentioned": "Suburbs",
        "property_types":    "Property Types",
        "timeline":          "Timeline",
        "sentiment":         "Sentiment",
        "engagement_score":  "Engagement",
        "ai_summary":        "Summary",
        "last_activity_at":  "Last Activity",
        "created_at":        "Created At",
    }

    col_widths = {
        "conversation_id":   38, "lead_name": 22, "lead_email": 28,
        "lead_phone": 18, "total_messages": 12, "status": 14,
        "lead_tier": 12, "lead_score": 10, "intent": 14,
        "budget_min": 14, "budget_max": 14, "budget_currency": 12,
        "suburbs_mentioned": 28, "property_types": 22, "timeline": 16,
        "sentiment": 14, "engagement_score": 14, "ai_summary": 40,
        "last_activity_at": 22, "created_at": 22,
    }

    DATE_FIELDS = {"last_activity_at", "created_at"}

    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=pretty_headers.get(field, field))
        cell.font      = _font(color=C["white"], bold=True, size=10)
        cell.fill      = _fill(C["brown"])
        cell.alignment = CENTER
        cell.border    = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 16)
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(flat_rows, start=2):
        row_fill = _fill(C["light"]) if row_idx % 2 == 0 else _fill(C["white"])
        for col_idx, field in enumerate(fieldnames, start=1):
            raw_val = row.get(field, "")
            val     = _fmt_date(raw_val) if field in DATE_FIELDS else raw_val

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = CELL_BORDER
            cell.alignment = WRAP if field == "ai_summary" else CENTER
            cell.font      = _font(color=C["dark"], size=9)

            tier_hex = TIER_ACCENT.get(val) if field == "lead_tier" else None
            if tier_hex:
                cell.fill = _fill(tier_hex)
                cell.font = _font(color=C["white"], bold=True, size=9)
            else:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 18

    # ── Sheet 2: Analytics ────────────────────────────────────────────────────
    wc = wb.create_sheet("Analytics")
    wc.sheet_view.showGridLines = False

    hdr_font = _font(color=C["white"], bold=True, size=11)
    lbl_font = _font(color=C["taupe"], bold=True, size=10)
    val_font = _font(color=C["dark"], size=10)

    def _analytics_header(cell_ref, label, fill_color=C["teal"]):
        cell = wc[cell_ref]
        cell.value     = label
        cell.font      = hdr_font
        cell.fill      = _fill(fill_color)
        cell.alignment = CENTER
        cell.border    = CELL_BORDER

    tier_counts  = Counter(r.get("lead_tier")  or "unknown" for r in flat_rows)
    tier_labels  = list(tier_counts.keys())
    tier_values  = list(tier_counts.values())

    wc.column_dimensions["A"].width = 20
    wc.column_dimensions["B"].width = 12

    _analytics_header("A1", "Tier",  fill_color=C["brown"])
    _analytics_header("B1", "Count", fill_color=C["brown"])
    for i, (label, count) in enumerate(zip(tier_labels, tier_values), start=2):
        accent = TIER_ACCENT.get(label, C["brownShade"])
        c1 = wc.cell(row=i, column=1, value=label)
        c1.fill = _fill(accent); c1.font = _font(color=C["white"], bold=True, size=10)
        c1.border = CELL_BORDER; c1.alignment = CENTER
        c2 = wc.cell(row=i, column=2, value=count)
        c2.fill = _fill(C["light"]); c2.font = val_font
        c2.border = CELL_BORDER; c2.alignment = CENTER

    sentiment_counts = Counter(r.get("sentiment") or "unknown" for r in flat_rows)
    sent_labels      = list(sentiment_counts.keys())
    sent_values      = list(sentiment_counts.values())

    wc.column_dimensions["D"].width = 20
    wc.column_dimensions["E"].width = 12

    _analytics_header("D1", "Sentiment", fill_color=C["teal"])
    _analytics_header("E1", "Count",     fill_color=C["teal"])
    for i, (label, count) in enumerate(zip(sent_labels, sent_values), start=2):
        sent_hex = SENT_FILL.get(label, C["light"])
        c1 = wc.cell(row=i, column=4, value=label)
        c1.fill = _fill(sent_hex)
        c1.font = _font(color=C["white"] if sent_hex != C["light"] else C["taupe"], bold=True, size=10)
        c1.border = CELL_BORDER; c1.alignment = CENTER
        c2 = wc.cell(row=i, column=5, value=count)
        c2.fill = _fill(C["light"]); c2.font = val_font
        c2.border = CELL_BORDER; c2.alignment = CENTER

    intent_counts = Counter(r.get("intent") or "unknown" for r in flat_rows)
    intent_labels = list(intent_counts.keys())
    intent_values = list(intent_counts.values())

    intent_start_row = 30
    _analytics_header(wc.cell(row=intent_start_row, column=1).coordinate, "AI Intent", fill_color=C["burntOrange"])
    _analytics_header(wc.cell(row=intent_start_row, column=2).coordinate, "Count",     fill_color=C["burntOrange"])
    for i, (label, count) in enumerate(zip(intent_labels, intent_values), start=intent_start_row + 1):
        c1 = wc.cell(row=i, column=1, value=label)
        c1.font = lbl_font; c1.fill = _fill(C["light"]); c1.border = CELL_BORDER; c1.alignment = CENTER
        c2 = wc.cell(row=i, column=2, value=count)
        c2.font = val_font;  c2.fill = _fill(C["white"]); c2.border = CELL_BORDER; c2.alignment = CENTER

    pie1 = PieChart()
    pie1.title  = "Lead Tier Distribution"
    pie1.style  = 10
    pie1.width  = 14
    pie1.height = 10
    pie1.add_data(Reference(wc, min_col=2, min_row=1, max_row=1 + len(tier_labels)), titles_from_data=True)
    pie1.set_categories(Reference(wc, min_col=1, min_row=2, max_row=1 + len(tier_labels)))
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showPercent  = True
    pie1.dataLabels.showCatName  = True
    _color_pie_slices(pie1, [TIER_ACCENT.get(l, C["brownShade"]) for l in tier_labels])
    wc.add_chart(pie1, "G1")

    pie2 = PieChart()
    pie2.title  = "Sentiment Distribution"
    pie2.style  = 10
    pie2.width  = 14
    pie2.height = 10
    pie2.add_data(Reference(wc, min_col=5, min_row=1, max_row=1 + len(sent_labels)), titles_from_data=True)
    pie2.set_categories(Reference(wc, min_col=4, min_row=2, max_row=1 + len(sent_labels)))
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showPercent  = True
    pie2.dataLabels.showCatName  = True
    _color_pie_slices(pie2, [SENT_FILL.get(l, C["brownShade"]) for l in sent_labels])
    wc.add_chart(pie2, "O1")

    bar = BarChart()
    bar.type         = "col"
    bar.title        = "Lead Intent Breakdown"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Intent"
    bar.style        = 10
    bar.width        = 28
    bar.height       = 10
    bar.add_data(Reference(wc, min_col=2, min_row=intent_start_row, max_row=intent_start_row + len(intent_labels)), titles_from_data=True)
    bar.set_categories(Reference(wc, min_col=1, min_row=intent_start_row + 1, max_row=intent_start_row + len(intent_labels)))
    BAR_COLORS = [C["brown"], C["teal"], C["burntOrange"], C["taupe"], C["warm"], C["cold"], C["nurture"]]
    _color_bar_series(bar, [BAR_COLORS[i % len(BAR_COLORS)] for i in range(len(intent_labels))])
    wc.add_chart(bar, "G30")

    # ── Sheet 3: Summary Stats ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    total      = len(flat_rows)
    leads_only = sum(1 for r in flat_rows if str(r.get("is_lead", "")).lower() in ("true", "1"))
    hot        = tier_counts.get("hot",     0)
    warm       = tier_counts.get("warm",    0)
    nurture    = tier_counts.get("nurture", 0)
    cold       = tier_counts.get("cold",    0)

    SECTION_FILLS = [_fill(C["brown"]), _fill(C["taupe"]), _fill(C["teal"]), _fill(C["burntOrange"])]

    def write_section(row, title, section_idx=0):
        fill = SECTION_FILLS[section_idx % len(SECTION_FILLS)]
        cell = ws2.cell(row=row, column=1, value=title)
        cell.fill      = fill
        cell.font      = _font(color=C["white"], bold=True, size=12)
        cell.alignment = CENTER
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws2.row_dimensions[row].height = 24
        ws2.cell(row=row, column=2).fill = fill

    def write_kv(row, label, value, label_color=C["taupe"], value_bg=C["light"]):
        lc = ws2.cell(row=row, column=1, value=label)
        lc.font = _font(color=label_color, bold=True, size=11)
        lc.fill = _fill(C["light"]); lc.border = CELL_BORDER; lc.alignment = WRAP
        vc = ws2.cell(row=row, column=2, value=value)
        vc.font = _font(color=C["dark"], size=11)
        vc.fill = _fill(value_bg); vc.border = CELL_BORDER; vc.alignment = CENTER
        ws2.row_dimensions[row].height = 20

    write_section(1, "Export Summary", 0)
    write_kv(2, "Total Conversations",  total)
    write_kv(3, "Total Leads Captured", leads_only)
    write_kv(4, "Conversion Rate", f"{round(leads_only / total * 100, 2) if total else 0}%")

    write_section(6, "Lead Tiers", 1)
    for i, (label, val, accent) in enumerate([
        ("Hot Leads",     hot,     C["hot"]),
        ("Warm Leads",    warm,    C["warm"]),
        ("Nurture Leads", nurture, C["nurture"]),
        ("Cold Leads",    cold,    C["cold"]),
    ], start=7):
        write_kv(i, label, val, label_color=C["taupe"], value_bg=accent)
        ws2.cell(row=i, column=2).font = _font(color=C["white"], bold=True, size=11)

    write_section(12, "Sentiment Breakdown", 2)
    for i, (sent, cnt) in enumerate(sentiment_counts.items(), start=13):
        bg = SENT_FILL.get(sent, C["light"])
        write_kv(i, sent.capitalize(), cnt, value_bg=bg)
        if bg != C["light"]:
            ws2.cell(row=i, column=2).font = _font(color=C["white"], bold=True, size=11)

    intent_section_row = 14 + len(sentiment_counts)
    write_section(intent_section_row, "Intent Breakdown", 3)
    for i, (intent, cnt) in enumerate(intent_counts.items(), start=intent_section_row + 1):
        write_kv(i, intent.capitalize(), cnt)

    STATUS_GUIDE = [
        ("in_progress", "Chat is active, visitor is currently messaging."),
        ("closed",      "Auto-closed after 15 min of inactivity by the hourly job."),
        ("summarized",  "AI has extracted lead insights from the closed conversation."),
        ("emailed",     "Summary sent to the visitor (if email provided) and internal team."),
        ("archived",    "Conversation is old and has been moved to long-term storage."),
    ]
    status_section_row = intent_section_row + len(intent_counts) + 3
    write_section(status_section_row, "Status Guide", 0)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 62

    for i, (st, desc) in enumerate(STATUS_GUIDE, start=status_section_row + 1):
        lc = ws2.cell(row=i, column=1, value=st)
        lc.font = _font(color=C["white"], bold=True, size=10)
        lc.fill = _fill(C["taupe"]); lc.border = CELL_BORDER; lc.alignment = CENTER
        ws2.row_dimensions[i].height = 20
        rc = ws2.cell(row=i, column=2, value=desc)
        rc.font = _font(color=C["dark"], size=10)
        rc.fill = _fill(C["light"]); rc.border = CELL_BORDER; rc.alignment = WRAP

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
